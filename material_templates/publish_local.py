#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
本機教材發布：讀 Excel（含 _Config / _Schema / _Publish / _Layout）
→ 產出 .meta.json / .script.txt / _manifest.json / _layout.json
不上傳 Drive、不轉 Google Sheets。

用法：
  ./publish_local.sh /path/to/workbook.xlsx
  ./publish_local.sh /path/to/workbook.xlsx --out /path/to/output_dir
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(
        "缺少 openpyxl。請在本目錄執行：\n"
        "  python3 -m venv .venv && .venv/bin/pip install -r requirements-publish.txt\n"
        "  .venv/bin/python publish_local.py <你的.xlsx>",
        file=sys.stderr,
    )
    sys.exit(1)


def col_letter_to_index(col: str) -> int:
    s = (col or "").strip().upper()
    if not s:
        return -1
    n = 0
    for ch in s:
        if not ("A" <= ch <= "Z"):
            return -1
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def is_truthy_yn(val) -> bool:
    s = str(val if val is not None else "").strip().upper()
    return s in ("Y", "YES", "1", "是")


def is_empty(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def rows_to_maps(headers_raw, data_rows) -> list[dict]:
    """
    共用轉換：第一列當表頭（轉小寫），其餘列轉成 {表頭: 值} 的 dict。
    供舊格式（單一活頁 iter_rows）與新格式（合併 _Setup 活頁切段）共用。
    """
    headers = [str(h or "").strip().lower() for h in (headers_raw or [])]
    out = []
    for row in data_rows:
        obj = {}
        empty = True
        for i, key in enumerate(headers):
            if not key:
                continue
            v = row[i] if i < len(row) else None
            if not is_empty(v):
                empty = False
            obj[key] = v
        if not empty:
            out.append(obj)
    return out


def sheet_as_maps(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"找不到活頁：{sheet_name}")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return []
    return rows_to_maps(headers_raw, rows_iter)


# 合併活頁（_Setup）的區段標記：一列裡第一格是 #_Config／#_Schema／#_Publish／#_Layout
# （前面的 # 可有可無、大小寫不拘），下一列當表頭，再往下讀到下一個標記或空白列為止。
SETUP_SHEET_NAME = "_Setup"
SETUP_SECTION_NAMES = ("_Config", "_Schema", "_Publish", "_Layout")
_SETUP_MARKER_RE = re.compile(r"^#?_?(config|schema|publish|layout)\s*$", re.IGNORECASE)


def _setup_marker_section(cell_value) -> str | None:
    text = str(cell_value or "").strip()
    m = _SETUP_MARKER_RE.match(text)
    if not m:
        return None
    return "_" + m.group(1).capitalize()


def find_setup_sections(wb) -> dict[str, list[dict]] | None:
    """
    偵測並解析合併格式的 `_Setup` 活頁：把 _Config／_Schema／_Publish／_Layout
    四小段從同一個活頁切出來，各自轉成跟舊格式 sheet_as_maps 一樣的 list[dict]。
    沒有 `_Setup` 活頁時回傳 None（呼叫端會改用舊的 4 分頁各自讀取）。
    """
    if SETUP_SHEET_NAME not in wb.sheetnames:
        return None
    matrix = load_sheet_matrix(wb[SETUP_SHEET_NAME])
    sections: dict[str, list[dict]] = {name: [] for name in SETUP_SECTION_NAMES}
    i = 0
    n = len(matrix)
    while i < n:
        row = matrix[i]
        marker = _setup_marker_section(row[0] if row else None)
        if marker is None:
            i += 1
            continue
        header_row = matrix[i + 1] if i + 1 < n else ()
        data_rows = []
        j = i + 2
        while j < n:
            row_j = matrix[j]
            if _setup_marker_section(row_j[0] if row_j else None) is not None:
                break
            data_rows.append(row_j)
            j += 1
        sections[marker] = rows_to_maps(header_row, data_rows)
        i = j
    return sections


def read_config(rows: list[dict]) -> dict:
    cfg = {"material_folder": "GEPT-2_vocab", "last_row_column": "A"}
    for r in rows:
        k = str(r.get("key") or "").strip()
        v = r.get("value")
        if k == "material_folder" and not is_empty(v):
            cfg["material_folder"] = str(v).strip()
        if k == "last_row_column" and not is_empty(v):
            cfg["last_row_column"] = str(v).strip()
    return cfg


def read_schemas(rows: list[dict]) -> dict:
    by_schema: dict[str, list] = {}
    for r in rows:
        sid = str(r.get("schema_id") or "").strip()
        if not sid:
            continue
        by_schema.setdefault(sid, []).append(
            {
                "semantic_key": str(r.get("semantic_key") or "").strip(),
                "excel_col": str(r.get("excel_col") or "").strip(),
                "send_to_ai": is_truthy_yn(r.get("send_to_ai")),
                "display": is_truthy_yn(r.get("display")),
            }
        )
    return by_schema


def read_publish_rules(rows: list[dict]) -> list[dict]:
    return [r for r in rows if is_truthy_yn(r.get("enabled"))]


def read_layout_profiles(rows: list[dict]) -> list[dict]:
    """
    讀 _Layout 段落／活頁：同一教材共用一份；列＝可選的排版／欄位組合。
    沒有列時回傳 []（不阻斷發布）。
    """
    profiles = []
    for r in rows:
        if not is_truthy_yn(r.get("enabled")):
            continue
        pid = str(r.get("profile_id") or "").strip()
        if not pid:
            continue
        lpp_raw = r.get("lines_per_page")
        try:
            lpp = int(lpp_raw) if lpp_raw is not None and str(lpp_raw).strip() != "" else 10
        except (TypeError, ValueError):
            lpp = 10
        if lpp <= 0:
            lpp = 10
        fields_answer = str(
            r.get("fields_answer") or r.get("answer_fields") or ""
        ).strip()
        profiles.append(
            {
                "profile_id": pid,
                "label": str(r.get("label") or pid).strip(),
                "fields": str(r.get("fields") or "").strip(),
                "fields_answer": fields_answer,
                # quiz_prompt／quiz_answer：專供「線上卷」單一提示／單一答案用，跟
                # fields／fields_answer（印刷多欄排版）分開；沒填時線上卷退回舊的
                # 「fields 第2欄＝提示、第3欄＝答案」慣例（相容 GEPT 句子翻譯教材）。
                # 見 .cursor/rules/material-publish-setup-format.mdc。
                "quiz_prompt": str(r.get("quiz_prompt") or "").strip(),
                "quiz_answer": str(r.get("quiz_answer") or "").strip(),
                "lines_per_page": lpp,
                "is_default": is_truthy_yn(r.get("is_default")),
                "note": str(r.get("note") or "").strip(),
            }
        )
    return profiles


def build_col_maps(schemas: dict) -> dict:
    """excel_col → semantic_key（依 schema_id）；供線上卷公式求值。"""
    out: dict[str, dict[str, str]] = {}
    for sid, fields in (schemas or {}).items():
        m: dict[str, str] = {}
        for f in fields or []:
            col = str(f.get("excel_col") or "").strip().upper()
            key = str(f.get("semantic_key") or "").strip()
            if col and key:
                m[col] = key
        if m:
            out[str(sid)] = m
    return out


def build_layout_payload(
    cfg: dict, profiles: list[dict], published_at: str, schemas: dict | None = None
) -> dict | None:
    if not profiles:
        return None
    default_id = ""
    for p in profiles:
        if p.get("is_default"):
            default_id = p["profile_id"]
            break
    if not default_id:
        default_id = profiles[0]["profile_id"]
    col_maps = build_col_maps(schemas or {})
    # 扁平 col_map：多 schema 時後寫覆蓋；通常一份教材共用一 schema
    flat: dict[str, str] = {}
    for _sid, m in col_maps.items():
        flat.update(m)
    return {
        "published_at": published_at,
        "material_folder": cfg.get("material_folder") or "",
        "default_profile_id": default_id,
        "col_map": flat,
        "col_maps": col_maps,
        "profiles": [
            {
                "profile_id": p["profile_id"],
                "label": p["label"],
                "fields": p["fields"],
                "fields_answer": p.get("fields_answer") or "",
                "quiz_prompt": p.get("quiz_prompt") or "",
                "quiz_answer": p.get("quiz_answer") or "",
                "lines_per_page": p["lines_per_page"],
                "note": p.get("note") or "",
            }
            for p in profiles
        ],
    }


def load_sheet_matrix(ws) -> list[tuple]:
    """一次讀完整張表（values_only），比逐格 ws.cell 快很多。"""
    return list(ws.iter_rows(values_only=True))


def get_last_data_row_from_matrix(matrix: list[tuple], col_letter: str) -> int:
    col = col_letter_to_index(col_letter)
    if col < 0:
        col = 0
    for r in range(len(matrix), 0, -1):
        row = matrix[r - 1]
        v = row[col] if col < len(row) else None
        if not is_empty(v):
            return r
    return 1


def build_rows_from_matrix(
    matrix: list[tuple],
    sheet_name: str,
    rule: dict,
    schema_fields: list,
    cfg: dict,
) -> list[dict]:
    try:
        start_row = int(rule.get("row_start") or 2)
    except (TypeError, ValueError):
        start_row = 2
    if start_row < 1:
        start_row = 2

    end_raw = str(rule.get("row_end") or "LAST").strip().upper()
    if end_raw == "LAST":
        end_row = get_last_data_row_from_matrix(matrix, cfg["last_row_column"])
    else:
        try:
            end_row = int(end_raw)
        except ValueError as e:
            raise SystemExit(f"列範圍無效：{sheet_name} ({start_row}-{end_raw})") from e

    if end_row < start_row:
        raise SystemExit(f"列範圍無效：{sheet_name} ({start_row}-{end_raw})")

    has_script = any(f.get("semantic_key") == "script" and f.get("excel_col") for f in schema_fields)
    if not has_script:
        raise SystemExit(f"schema 缺少 script 欄位映射（schema_id={rule.get('schema_id')}）")

    field_map = []
    for f in schema_fields:
        key = f.get("semantic_key") or ""
        col_letter = f.get("excel_col") or ""
        if not key or not col_letter:
            continue
        ci = col_letter_to_index(col_letter)
        if ci < 0:
            continue
        field_map.append((key, ci))

    out = []
    formula_warned = False
    end_row = min(end_row, len(matrix))
    for r in range(start_row, end_row + 1):
        row = matrix[r - 1]
        row_obj = {"_source_row": r}
        has_script_val = False
        for key, ci in field_map:
            raw = row[ci] if ci < len(row) else None
            if isinstance(raw, str) and raw.startswith("="):
                if not formula_warned:
                    print(
                        "警告：偵測到尚未快取結果的公式。請用 Excel 開啟存檔一次後再跑。",
                        file=sys.stderr,
                    )
                    formula_warned = True
                continue
            if is_empty(raw):
                continue
            row_obj[key] = raw
            if key == "script":
                has_script_val = True
        if has_script_val:
            out.append(row_obj)
    return out


def safe_filename(name: str) -> str:
    name = str(name or "").strip()
    name = re.sub(r'[\\/:*?"<>|]', "_", name)
    return name or "output"


def json_safe(v):
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return v


def publish(xlsx_path: Path, out_root: Path | None) -> Path:
    print(f"讀取：{xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    setup_sections = find_setup_sections(wb)
    if setup_sections is not None:
        print(f"偵測到合併活頁 `{SETUP_SHEET_NAME}`（新格式）")
        cfg = read_config(setup_sections["_Config"])
        schemas = read_schemas(setup_sections["_Schema"])
        rules = read_publish_rules(setup_sections["_Publish"])
        layout_profiles = read_layout_profiles(setup_sections["_Layout"])
    else:
        print("偵測到傳統 4 分頁格式（_Config／_Schema／_Publish／_Layout 各自獨立）")
        cfg = read_config(sheet_as_maps(wb, "_Config"))
        schemas = read_schemas(sheet_as_maps(wb, "_Schema"))
        rules = read_publish_rules(sheet_as_maps(wb, "_Publish"))
        layout_profiles = read_layout_profiles(
            sheet_as_maps(wb, "_Layout") if "_Layout" in wb.sheetnames else []
        )
    if not rules:
        raise SystemExit("_Publish 沒有 enabled=Y 的規則（請把要發布的列改成 Y）")

    out_dir = out_root or (xlsx_path.parent / "published" / cfg["material_folder"])
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"輸出目錄：{out_dir.resolve()}")
    print(f"共 {len(rules)} 條 enabled=Y 規則")
    if layout_profiles:
        print(f"共 {len(layout_profiles)} 個 _Layout 排版（同教材共用）")
    else:
        print("（無 _Layout 或沒有 enabled=Y → 略過 _layout.json）")
    print()

    outputs = []
    now = datetime.now(timezone.utc).isoformat()
    matrix_cache: dict[str, list] = {}

    for i, rule in enumerate(rules, 1):
        sid = str(rule.get("schema_id") or "").strip()
        fields = schemas.get(sid)
        if not fields:
            raise SystemExit(f"找不到 schema_id：{sid}")

        sheet_name = str(rule.get("source_sheet") or "").strip()
        if not sheet_name:
            raise SystemExit("_Publish 缺少 source_sheet")
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"找不到來源活頁：{sheet_name}")

        print(f"[{i}/{len(rules)}] 處理活頁 {sheet_name} …", flush=True)
        if sheet_name not in matrix_cache:
            matrix_cache[sheet_name] = load_sheet_matrix(wb[sheet_name])
        rows = build_rows_from_matrix(matrix_cache[sheet_name], sheet_name, rule, fields, cfg)
        if not rows:
            print(f"  警告：{sheet_name} 沒有含 script 的列（請檢查 _Schema 欄位與公式快取）", file=sys.stderr)

        clean_rows = []
        for row in rows:
            copy = {k: json_safe(v) for k, v in row.items() if not str(k).startswith("_")}
            clean_rows.append(copy)

        meta_name = str(rule.get("output_meta") or "").strip()
        txt_name = str(rule.get("output_txt") or "").strip()
        src_label = sheet_name or "sheet"
        if not meta_name:
            meta_name = f"{src_label}.meta.json"
        meta_name = safe_filename(meta_name)
        if txt_name:
            txt_name = safe_filename(txt_name)

        meta_path = out_dir / meta_name
        meta_path.write_text(
            json.dumps(clean_rows, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(f"  ✓ {meta_name}（{len(clean_rows)} 列）", flush=True)

        if txt_name:
            lines = [str(row.get("script") or "").strip() for row in rows]
            lines = [ln for ln in lines if ln]
            (out_dir / txt_name).write_text(
                "\n".join(lines) + ("\n" if lines else ""),
                encoding="utf-8",
            )
            print(f"  ✓ {txt_name}（{len(lines)} 行）", flush=True)

        outputs.append(
            {
                "source_sheet": sheet_name,
                "meta": meta_name,
                "txt": txt_name or None,
                "rowCount": len(clean_rows),
                "schema_id": sid,
            }
        )

    layout_payload = build_layout_payload(cfg, layout_profiles, now, schemas)
    if layout_payload:
        (out_dir / "_layout.json").write_text(
            json.dumps(layout_payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(f"\n✓ _layout.json（預設 {layout_payload['default_profile_id']}，"
              f"{len(layout_payload['profiles'])} 種可選）")

    manifest = {
        "published_at": now,
        "source_file": str(xlsx_path.name),
        "material_folder": cfg["material_folder"],
        "root_kind": "local",
        "outputs": outputs,
        "layout": "_layout.json" if layout_payload else None,
    }
    (out_dir / "_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"✓ _manifest.json")
    print(f"完成：{out_dir.resolve()}")
    wb.close()
    return out_dir


def main():
    parser = argparse.ArgumentParser(description="本機發布 Excel → meta.json / script.txt")
    parser.add_argument("xlsx", type=Path, help="本機 Excel 檔路徑（.xlsx）")
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="輸出目錄（預設：xlsx 同層 published/<material_folder>/）",
    )
    args = parser.parse_args()
    xlsx = args.xlsx.expanduser().resolve()
    if not xlsx.is_file():
        raise SystemExit(f"找不到檔案：{xlsx}")
    if xlsx.suffix.lower() not in (".xlsx", ".xlsm"):
        raise SystemExit("請使用 .xlsx（不支援舊版 .xls）")
    publish(xlsx, args.out.expanduser().resolve() if args.out else None)


if __name__ == "__main__":
    main()
