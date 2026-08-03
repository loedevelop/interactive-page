#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把舊格式教材 Excel（_Config／_Schema／_Publish／_Layout 四個分頁各自獨立）
轉成新格式（合併成一個 `_Setup` 分頁，用 #_Config 等標記分段）。

其餘活頁（內容活頁 A~Z、V… 等）原樣保留，不動、不重新計算公式。
輸出成新檔（不覆蓋原檔），你可以先打開檢查沒問題再取代。

用法：
  ./convert_to_setup.sh /path/to/old_workbook.xlsx
  ./convert_to_setup.sh /path/to/old_workbook.xlsx --out /path/to/new_workbook.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(
        "缺少 openpyxl。請在本目錄執行：\n"
        "  python3 -m venv .venv && .venv/bin/pip install -r requirements-publish.txt\n"
        "  .venv/bin/python convert_to_setup.py <你的.xlsx>",
        file=sys.stderr,
    )
    sys.exit(1)

from publish_local import SETUP_SECTION_NAMES, SETUP_SHEET_NAME, load_sheet_matrix


def _strip_trailing_empty_rows(rows: list[tuple]) -> list[tuple]:
    out = list(rows)
    while out and all(v is None or (isinstance(v, str) and v.strip() == "") for v in out[-1]):
        out.pop()
    return out


def convert(xlsx_path: Path, out_path: Path) -> Path:
    print(f"讀取（舊格式）：{xlsx_path}")
    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    section_rows: dict[str, list[tuple]] = {}
    for name in SETUP_SECTION_NAMES:
        if name not in wb_values.sheetnames:
            print(f"  （沒有 {name} 分頁，略過）")
            continue
        rows = _strip_trailing_empty_rows(load_sheet_matrix(wb_values[name]))
        section_rows[name] = rows
        print(f"  讀到 {name}：{len(rows)} 列（含表頭）")
    wb_values.close()

    if "_Config" not in section_rows:
        raise SystemExit("舊檔缺少 _Config 分頁，無法轉換（請確認這是本教材發布格式的 Excel）")

    print(f"開啟（保留其他活頁與公式）：{xlsx_path}")
    wb_out = openpyxl.load_workbook(xlsx_path)

    for name in SETUP_SECTION_NAMES:
        if name in wb_out.sheetnames:
            del wb_out[name]

    ws = wb_out.create_sheet(SETUP_SHEET_NAME, 0)
    for name in SETUP_SECTION_NAMES:
        rows = section_rows.get(name)
        if not rows:
            continue
        ws.append((f"#{name}",))
        for row in rows:
            ws.append(list(row))
        ws.append(())  # 空白列分隔，方便閱讀

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(out_path)
    wb_out.close()
    print(f"✓ 已輸出新格式：{out_path.resolve()}")
    print("請打開確認 `_Setup` 分頁內容正確、其他活頁（內容／公式）沒有跑掉，再決定是否取代原檔。")
    return out_path


def main():
    import argparse

    parser = argparse.ArgumentParser(description="把舊格式（4 分頁）教材 Excel 轉成新格式（合併 _Setup 分頁）")
    parser.add_argument("xlsx", type=Path, help="舊格式教材 Excel 路徑（.xlsx）")
    parser.add_argument("--out", type=Path, default=None, help="輸出檔路徑（預設：同層 <檔名>.setup.xlsx）")
    args = parser.parse_args()

    xlsx = args.xlsx.expanduser().resolve()
    if not xlsx.is_file():
        raise SystemExit(f"找不到檔案：{xlsx}")
    if xlsx.suffix.lower() not in (".xlsx", ".xlsm"):
        raise SystemExit("請使用 .xlsx（不支援舊版 .xls）")

    if args.out:
        out_path = args.out.expanduser().resolve()
    else:
        out_path = xlsx.with_suffix("")
        out_path = out_path.parent / f"{out_path.name}.setup.xlsx"

    convert(xlsx, out_path)


if __name__ == "__main__":
    main()
